# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import datetime

# useful for handling different item types with a single interface
from itemadapter import ItemAdapter


# class MydemoPipeline:
#     def process_item(self, item, spider):
#         return item

import openpyxl
from mydemo.items import DoubanItem, PixivItem, PixivDownloadItem, HouseItem, BiliItem, NetItem, NetWordItem
import os
import requests
from scrapy import Request
from scrapy.exceptions import DropItem
from scrapy.pipelines.images import ImagesPipeline


class DoubanItemPipeline:

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = 'Top250'
        self.sheet.append(('名称', '评分', '名言'))

    def process_item(self, item: DoubanItem, spider):
        self.sheet.append((item['title'], item['score'], item['motto']))
        return item

    def close_spider(self, spider):
        # 获取根目录路径
        root_path = os.path.abspath(os.path.dirname(__file__))
        self.wb.save(f'{root_path}/output/豆瓣电影数据.xlsx')


class PixivPipeline:

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = 'weekly'
        self.sheet.append(('标题', '作者', 'P_ID', "链接"))

    def process_item(self, item: PixivItem, spider):
        self.sheet.append((item['title'], item['user_name'], item['p_id'], item['re_url']))
        return item

    def close_spider(self, spider):
        # 获取根目录路径
        root_path = os.path.abspath(os.path.dirname(__file__))
        self.wb.save(f'{root_path}/output/pixiv_weekly_rank数据.xlsx')


class PixivDownloadPipeline:
    def __init__(self):
        self.root_path = os.path.abspath(os.path.dirname(__file__))

    def process_item(self, item: PixivDownloadItem, spider):
        """对item进行处理，如果是多p就新建文件夹放进去"""
        if item["is_many"]:
            save_path = f"{self.root_path}/output/{item['folder_name']}/"
        else:
            save_path = f"{self.root_path}/output/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        for img in item["final_urls"]:
            file_name = f"{img['title']}.{img['file_type']}"
            resp = requests.get(img["url"], headers=item["headers"])
            with open(f"{save_path}{file_name}", 'wb') as file:
                file.write(resp.content)
        return item


class PixivImagePipeline(ImagesPipeline):
    root_path = os.path.abspath(os.path.dirname(__file__))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs )
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "下载结果"
        self.worksheet.append(('文件夹', '文件名', '下载状态', '下载链接'))

    # def open_spider(self, spider):

    def get_media_requests(self, item: PixivDownloadItem, info):
        for img in item["final_urls"]:
            url = img["url"]
            headers = item["headers"]
            file_name = f"{img['title']}.{img['file_type']}"
            if item["is_many"]:
                save_path = f"{PixivImagePipeline.root_path}/output/{item['folder_name']}"
            else:
                save_path = f"{PixivImagePipeline.root_path}/output"
            yield Request(url, headers=headers, meta={"file_name": file_name,
                                                      "save_path": save_path
                                                      })

    def file_path(self, request, response=None, info=None, *, item=None):
        """返回保存的完整路径，例如static/output/aaa.jpg"""
        file_name = request.meta["file_name"]
        save_path = request.meta["save_path"]
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        filename = f'{save_path}/{file_name}'  # 构建图片保存路径和名称
        return filename

    def item_completed(self, results, item, info):
        """对下载完成的图片进行处理，这里的item是单个的，
        这里results有多个，是因为一个item中可能会有多个下载链接。
        每次对单个item中的所有下载链接处理完，就会调用一次item_completed"""
        for is_ok, result in results:
            # results是一个元组，(下载是否成功，dic）
            if is_ok:
                _status = "下载成功"
            else:
                _status = "下载失败"
                raise DropItem('Image Downloaded Failed')
            for img in item["final_urls"]:
                row = [item["folder_name"], img["title"], _status, result["url"]]
                self.worksheet.append(row)
        return item

    def close_spider(self, spider):
        # 获取根目录路径
        root_path = os.path.abspath(os.path.dirname(__file__))
        self.workbook.save(f'{root_path}/output/pixiv_weekly_下载情况.xlsx')


class ShellItemPipeline:
    root_path = os.path.abspath(os.path.dirname(__file__))

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = '房价数据'
        self.sheet.append(('城市', '区', '街道', '小区', '房屋信息', '总价', '单价'))

    def process_item(self, item: HouseItem, spider):
        self.sheet.append((item['house_city'], item['house_area'], item['house_street'],
                           item['house_community'], item['house_info'], item['house_total'],
                           item['house_unit']))
        return item

    def close_spider(self, spider):
        # 获取根目录路径
        date_string = datetime.datetime.today().strftime('%Y-%m-%d')
        save_path = f"{self.root_path}/output/{date_string}"
        city = spider.city_cn
        file_name = f"{city}_{date_string}.xlsx"
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        self.wb.save(f'{save_path}/{file_name}')


class BiliItemPipeline:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = 'bilibili'
        self.sheet.append(('视频标题', 'up主', '点击量', '评论数', '发布时间', '点赞', '收藏', '投币', '转发'))

    def process_item(self, item: BiliItem, spider):
        self.sheet.append((item["bili_title"], item['bili_author'], item['bili_clicks'],
                           item["bili_comments"], item["bili_uptime"], item["bili_likes"],
                           item["bili_favorites"], item["bili_coins"], item["bili_shares"]))
        return item

    def close_spider(self, spider):
        # 获取根目录路径
        root_path = os.path.abspath(os.path.dirname(__file__))
        self.wb.save(f'{root_path}/output/bilibili_data.xlsx')


class NetItemPipeline:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = '网易云'
        self.sheet.append(('歌曲名', '歌手名', '专辑名', '歌曲url'))

    def process_item(self, item: NetItem, spider):
        self.sheet.append((item["song_name"], item['singer_name'], item['cd_name'], item["song_url"]))
        return item

    def close_spider(self, spider):
        # 获取根目录路径
        root_path = os.path.abspath(os.path.dirname(__file__))
        self.wb.save(f'{root_path}/output/net_music_data.xlsx')


class NetWordItemPipeline:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = '网易云热评'
        self.sheet.append(('歌曲名', '用户nickname', '评论内容', '评论时间', '点赞数'))

    def process_item(self, item: NetWordItem, spider):
        self.sheet.append((item["song_name"], item['user_name'], item['content'], item["comment_date"],
                           item["liked_count"]))
        return item

    def close_spider(self, spider):
        root_path = os.path.abspath(os.path.dirname(__file__))
        self.wb.save(f'{root_path}/output/net_hotComments.xlsx')
